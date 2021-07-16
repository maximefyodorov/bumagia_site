# -*- coding: utf-8 -*-

import jinja2
import openpyxl

wb = openpyxl.load_workbook('w:/home/bumagia/www/templates/bumagia_data.xlsx', data_only=True)
templateLoader = jinja2.FileSystemLoader(searchpath='w:/home/bumagia/www/templates/All_templates/Inner_templates')
templateEnv = jinja2.Environment(loader=templateLoader)

ws = wb['common_data']

menu_data = []
for i in range (2, ws.max_column):
    menu_item = {}
    menu_item['link'] = ws.cell(row = 2, column = i).value
    menu_item['name'] = ws.cell(row = 9, column = i).value
    menu_item['prefix'] = ''
    menu_data.append(menu_item)

for i in range (2, ws.max_column):
    if (ws.cell(row = 1, column = i).value) == 1:
        my_product = ws.cell(row = 2, column = i).value
        page_data = {}
        for j in range (2, ws.max_row+1):
            page_data[ws.cell(row = j, column = 1).value] = ws.cell(row = j, column = i).value

        ws = wb[my_product]
        product_data = []
        for i in range (2, ws.max_row+1):
            row = {}
            for j in range (1, ws.max_column+1):
                row[ws.cell(row = 1, column = j).value] = ws.cell(row = i, column = j).value
            if row['show'] !=0:
                product_data.append(row)

        tmplt = templateEnv.get_template('{}_template.html'.format(my_product))
        with open('w:/home/bumagia/www/products/{}.html'.format(my_product), 'wb') as dest:
            output = tmplt.render(
                menu_data = menu_data,
                page_data = page_data,
                product_data = product_data
            )
            dest.write(output.encode('utf-8'))