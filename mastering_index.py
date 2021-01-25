# -*- coding: utf-8 -*-

import jinja2
import openpyxl

wb = openpyxl.load_workbook('w:/home/bumagia/www/templates/bumagia_data_v2.xlsx', data_only=True)
templateLoader = jinja2.FileSystemLoader(searchpath='w:/home/bumagia/www/templates/')
templateEnv = jinja2.Environment(loader=templateLoader)

ws = wb['index']
catalog_link = ws.cell(row = 1, column = 2).value
menu_data = []
for i in range (2, ws.max_column):
    menu_item = {}
    for j in range(2,9):
        menu_item[ws.cell(row = j, column = 1).value] = ws.cell(row = j, column = i).value
    menu_data.append(menu_item)

ws = wb['common_data']
page_data = {}
for i in range(2, ws.max_row+1):
    page_data[ws.cell(row = i, column = 1).value] = ws.cell(row = i, column = 2).value

ws = wb[page_data['section_name']]
product_data = []
for i in range (2, ws.max_row+1):
    row = {}
    for j in range (1, ws.max_column+1):
        row[ws.cell(row = 1, column = j).value] = ws.cell(row = i, column = j).value
    if row['main'] ==1:
        product_data.append(row)

tmplt = templateEnv.get_template('index_template_v2.html')
with open('w:/home/bumagia/www/index_v2.html', 'wb') as dest:
    output = tmplt.render(
        catalog_link = catalog_link,
        page_data = page_data,
        menu_data = menu_data,
        product_data = product_data
    )
    dest.write(output.encode('utf-8'))




# from jinja2 import Template
# import openpyxl

# wb = openpyxl.load_workbook('w:/home/bumagia/www/templates/bumagia_data_v2.xlsx', data_only=True)

# products_list = wb.sheetnames[1:]
# number_of_products = len(products_list)

# ws = wb['common_data']

# products_data = {}

# for i in range (number_of_products):
#     temp_dict = {}
#     for j in range (2, ws.max_row+1):
#         temp_dict[ws.cell(row = j, column = 1).value] = ws.cell(row = j, column = i+2).value
#     products_data[products_list[i]] = temp_dict


# my_product = 'ecofiller'

# ws = wb[my_product]

# my_arr_name = []
# my_arr_articul = []
# my_arr_full_articul = []
# my_arr_mame_and_articul = []
# my_arr_buylink = []

# my_len = 0

# for i in range(1, ws.max_row + 1):
#     # Управляющая колонка 1 (0 - игнорировать, 1 - выводить в общий список, 2 - выводить на витрину на главной)
#     control = ws.cell(row = i, column = 1).value 
#     if (control == 2):
#         my_len+=1
#         # Название (колонка 2)
#         my_arr_name.append(ws.cell(row = i, column = 2).value)
#         # Артикул без буквенной части - последние четыре цифры штрихкода (колонка 3)
#         my_arr_articul.append(str(ws.cell(row = i, column = 3).value)[9:])
#         # Полный артикул, индекс и цифры (колонка 4)
#         my_arr_full_articul.append(ws.cell(row = i, column = 4).value)
#         # Имя и полный артикул вместе (колонка 5)
#         my_arr_mame_and_articul.append(ws.cell(row = i, column = 5).value)
#         # Ссылка на покупку (колонка 6)
#         my_arr_buylink.append(ws.cell(row = i, column = 6).value)

# html = open('w:/home/bumagia/www/templates/index_template_v2.html', encoding='utf-8').read()
# tmplt = Template(html)
# with open('w:/home/bumagia/www/index_v2.html', 'wb') as dest:
#     output = tmplt.render(
#         p_list = products_list,
#         p_data = products_data,
#         p_num = number_of_products,
#         c_prod = my_product, 
#         v_num = my_len,
#         articul_list = my_arr_articul,
#         name_list = my_arr_name,
#         full_articul_list = my_arr_full_articul,
#         fullnameandarticul_list = my_arr_mame_and_articul,
#         buylink_list = my_arr_buylink
#     )
#     dest.write(output.encode('utf-8'))
