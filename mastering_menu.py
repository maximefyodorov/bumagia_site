# -*- coding: utf-8 -*-

import jinja2
import openpyxl

wb = openpyxl.load_workbook('h:/WebServers/home/bumagia/www/templates/bumagia_data.xlsx', data_only=True)

products_list = wb.sheetnames[1:]
number_of_products = len(products_list)

ws = wb['common_data']

my_menuitem_name =[]

templates_list = [
    'bookmarks',
    'modul_origami',
    'papercubes',
    'puzzles',
    'quilling',
    'quilling3d',
    'quilmagnet',
    'sketchbooks',
    'twosidespaper',
    'sertification',
    'instructions',
    'videolessons' 
]

link_prefix_list = [
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    'products/',
    '../products/',
    '../products/' 
]

path_prefix_list = [
    'products/',
    'products/',
    'products/',
    'products/',
    'products/',
    'products/',
    'products/',
    'products/',
    'products/',
    '',
    'learn/',
    'learn/' 
]


for i in range (2, ws.max_row+1):
    my_menuitem_name.append(ws.cell(row = 8, column = i).value)

templateLoader = jinja2.FileSystemLoader(searchpath='h:/WebServers/home/bumagia/www/templates/Inner_templates/')
templateEnv = jinja2.Environment(loader=templateLoader)
for i in (range(len(templates_list))):
    tmplt = templateEnv.get_template('{}_template.html'.format(templates_list[i]))
    with open('h:/WebServers/home/bumagia/www/{}{}.html'.format(path_prefix_list[i], templates_list[i]), 'wb') as dest:
        output = tmplt.render(
            p_list = products_list,
            p_num = number_of_products,
            menuname_list = my_menuitem_name,
            prefix = link_prefix_list[i]
        )
        dest.write(output.encode('utf-8'))